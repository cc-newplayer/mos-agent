from openpyxl import load_workbook
from pathlib import Path
from config import EXCEL_FILE


def _get_case_type_and_instruction(case):
    """根据用例信息推断测试类型和指令"""
    desc = str(case.get("description", ""))
    module = str(case.get("module", ""))

    # 闭环测试配置
    closedloop_configs = {
        "油管内容转写": {
            "type": "closedloop",
            "instruction": "对事件源中关注的账号最新的一篇帖子执行油管内容转写闭环"
        },
        "热点新闻蹭热度": {
            "type": "closedloop",
            "instruction": "用一个引流号找一个符合其人设的新闻热点执行蹭热度闭环"
        },
        "KOL跟评": {
            "type": "closedloop",
            "instruction": "用一个引流号对elonmusk最新的一篇帖子执行KOL跟评闭环"
        },
        "粉丝挖角": {
            "type": "closedloop",
            "instruction": "对事件源中关注的账号最新的一篇帖子执行粉丝挖角闭环"
        },
        "X投票裂变": {
            "type": "closedloop",
            "instruction": "对事件源中关注的账号最新的一篇帖子执行X投票裂变引爆闭环"
        },
    }

    # 菜单操作配置
    menu_operation_configs = {
        "增长闭环": {
            "type": "menu_operation",
            "operation": "growth_loop"
        },
        "定时任务": {
            "type": "menu_operation",
            "operation": "schedule_task"
        },
    }

    # 检查闭环测试
    for keyword, config in closedloop_configs.items():
        if keyword in desc:
            return config

    # 检查菜单操作
    for keyword, config in menu_operation_configs.items():
        if keyword in desc:
            return config

    # 默认：普通导航测试
    return {"type": "default"}


def load_test_cases(excel_path=None):
    """读取 Excel 测试用例，返回结构化列表"""
    if excel_path is None:
        excel_path = EXCEL_FILE

    if not Path(excel_path).exists():
        raise FileNotFoundError(f"测试用例文件不存在: {excel_path}")

    wb = load_workbook(excel_path)
    ws = wb.active

    cases = []
    headers = {}

    # 第一行是表头
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[cell.value] = col_idx

    # 从第二行开始读取用例
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
        if not headers:
            continue

        case_id = row[headers.get("用例编号", 1) - 1].value
        if not case_id:
            continue

        case = {
            "case_id": case_id,
            "description": row[headers.get("用例描述", 2) - 1].value if "用例描述" in headers else None,
            "module": row[headers.get("所属模块", 3) - 1].value if "所属模块" in headers else None,
            "level": row[headers.get("用例分级", 4) - 1].value if "用例分级" in headers else None,
            "precondition": row[headers.get("前置条件", 5) - 1].value if "前置条件" in headers else None,
            "steps": row[headers.get("步骤描述", 6) - 1].value if "步骤描述" in headers else None,
            "expected": row[headers.get("预期结果", 7) - 1].value if "预期结果" in headers else None,
            "test_data": row[headers.get("测试数据", 8) - 1].value if "测试数据" in headers else None,
        }

        # 添加测试类型和指令配置
        case.update(_get_case_type_and_instruction(case))

        cases.append(case)

    return cases


if __name__ == "__main__":
    cases = load_test_cases()
    print(f"加载了 {len(cases)} 条用例")
    for case in cases[:3]:
        print(case)
