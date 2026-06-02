#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
多闭环批量测试运行器
读取 Excel 文件，逐个执行闭环测试，汇总结果发送到飞书
"""
import asyncio
import sys
import io
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
import json
import requests

# 修复编码问题
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

sys.path.insert(0, str(Path(__file__).parent))

from executor.runner import run_test_case
from trace.tracker import get_trace_manager
from config import FEISHU_WEBHOOK


class MultiLoopRunner:
    """多闭环批量测试运行器"""

    def __init__(self, excel_path):
        self.excel_path = Path(excel_path)
        self.results = []
        self.start_time = None
        self.end_time = None

    def load_test_cases(self):
        """从 Excel 读取闭环测试用例"""
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {self.excel_path}")

        wb = load_workbook(self.excel_path)
        ws = wb.active

        cases = []

        # 读取表头
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value] = col_idx

        # 读取数据行
        for row_idx in range(2, ws.max_row + 1):
            case_id = ws.cell(row_idx, headers.get("闭环编号", 1)).value
            if not case_id:
                continue

            loop_name = ws.cell(row_idx, headers.get("闭环名称", 2)).value
            instruction = ws.cell(row_idx, headers.get("指令", 3)).value

            # 构造 case 对象，参考原有用例的结构
            case = {
                "case_id": case_id,
                "description": loop_name,
                "module": "决策对话",
                "level": "P1",
                "precondition": None,
                "steps": None,
                "expected": None,
                "test_data": None,
                "type": "closedloop",
                "instruction": instruction,
            }
            cases.append(case)

        return cases

    async def run_single_loop(self, case):
        """运行单个闭环测试"""
        print(f"\n{'='*70}")
        print(f"执行闭环: {case['case_id']} - {case['description']}")
        print(f"指令: {case['instruction']}")
        print(f"{'='*70}\n")

        try:
            result = await run_test_case(case)

            return {
                "case_id": case['case_id'],
                "name": case['description'],
                "instruction": case['instruction'],
                "status": result['status'],
                "error_msg": result.get('error_msg', ''),
                "duration": result.get('duration', 0),
                "trace_id": result.get('trace_id', 'N/A'),
                "timestamp": datetime.now().isoformat()
            }
        except Exception as e:
            return {
                "case_id": case['case_id'],
                "name": case['description'],
                "instruction": case['instruction'],
                "status": "ERROR",
                "error_msg": str(e),
                "duration": 0,
                "trace_id": "N/A",
                "timestamp": datetime.now().isoformat()
            }

    async def run_all_loops(self, cases):
        """逐个执行所有闭环测试（串行）"""
        self.start_time = datetime.now()

        for case in cases:
            result = await self.run_single_loop(case)
            self.results.append(result)

            # 打印结果
            status_icon = "✅" if result['status'] == 'PASS' else "❌" if result['status'] == 'FAIL' else "⚠️"
            print(f"\n{status_icon} 闭环 {result['case_id']} 结果: {result['status']}")
            if result['error_msg']:
                print(f"   错误: {result['error_msg']}")
            print(f"   耗时: {result['duration']}s")
            print(f"   追踪ID: {result['trace_id']}")

        self.end_time = datetime.now()

    def generate_summary(self):
        """生成测试汇总"""
        total = len(self.results)
        passed = sum(1 for r in self.results if r['status'] == 'PASS')
        failed = sum(1 for r in self.results if r['status'] == 'FAIL')
        error = sum(1 for r in self.results if r['status'] == 'ERROR')

        duration = (self.end_time - self.start_time).total_seconds() if self.end_time and self.start_time else 0

        return {
            "total": total,
            "passed": passed,
            "failed": failed,
            "error": error,
            "pass_rate": f"{(passed/total*100):.1f}%" if total > 0 else "0%",
            "duration": duration,
            "start_time": self.start_time.isoformat() if self.start_time else "",
            "end_time": self.end_time.isoformat() if self.end_time else "",
        }

    def send_to_feishu(self):
        """发送测试结果到飞书"""
        summary = self.generate_summary()

        # 构建飞书消息
        details = "\n".join([
            f"• 闭环 {r['case_id']} ({r['name']}): {r['status']}"
            for r in self.results
        ])

        message = {
            "msg_type": "text",
            "content": {
                "text": f"""🤖 多闭环测试完成

📊 测试汇总:
• 总数: {summary['total']}
• 通过: {summary['passed']} ✅
• 失败: {summary['failed']} ❌
• 错误: {summary['error']} ⚠️
• 通过率: {summary['pass_rate']}
• 总耗时: {summary['duration']:.1f}s

📋 详细结果:
{details}

⏰ 开始时间: {summary['start_time']}
⏰ 结束时间: {summary['end_time']}
"""
            }
        }

        try:
            response = requests.post(FEISHU_WEBHOOK, json=message)
            if response.status_code == 200:
                print("\n✅ 测试结果已发送到飞书")
            else:
                print(f"\n❌ 发送飞书失败: {response.status_code}")
        except Exception as e:
            print(f"\n❌ 发送飞书异常: {str(e)}")

    def save_results(self):
        """保存测试结果到 JSON 文件"""
        summary = self.generate_summary()

        output = {
            "summary": summary,
            "results": self.results
        }

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = Path(__file__).parent / f"multi_loop_results_{timestamp}.json"

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False, indent=2)

        print(f"\n✅ 测试结果已保存到: {output_file}")
        return output_file


async def main():
    excel_path = Path(__file__).parent.parent / "多闭环测试.xlsx"

    print("="*70)
    print("多闭环批量测试运行器")
    print("="*70)
    print(f"Excel 文件: {excel_path}\n")

    runner = MultiLoopRunner(excel_path)

    try:
        # 加载测试用例
        cases = runner.load_test_cases()
        print(f"✅ 加载了 {len(cases)} 条闭环测试用例\n")

        for case in cases:
            print(f"  • {case['case_id']}: {case['description']}")

        # 执行所有闭环
        print(f"\n{'='*70}")
        print("开始执行闭环测试...")
        print(f"{'='*70}")

        await runner.run_all_loops(cases)

        # 保存结果
        runner.save_results()

        # 发送到飞书
        runner.send_to_feishu()

        # 打印最终汇总
        summary = runner.generate_summary()
        print(f"\n{'='*70}")
        print("测试完成汇总")
        print(f"{'='*70}")
        print(f"总数: {summary['total']}")
        print(f"通过: {summary['passed']} ✅")
        print(f"失败: {summary['failed']} ❌")
        print(f"错误: {summary['error']} ⚠️")
        print(f"通过率: {summary['pass_rate']}")
        print(f"总耗时: {summary['duration']:.1f}s")

    except Exception as e:
        print(f"❌ 异常: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    asyncio.run(main())
