#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量加载器 - 支持 Excel、JSON、CSV 格式
"""
import json
import csv
from pathlib import Path


class BatchLoader:
    """批量加载器 - 自动检测文件格式并加载"""

    @staticmethod
    def load(file_path):
        """
        自动检测文件格式并加载

        Args:
            file_path: 文件路径（.xlsx/.json/.csv）

        Returns:
            loops: 闭环列表 [{'case_id': 1, 'name': '...', 'instruction': '...'}, ...]
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        if file_path.suffix.lower() == '.xlsx':
            return BatchLoader.load_excel(file_path)
        elif file_path.suffix.lower() == '.json':
            return BatchLoader.load_json(file_path)
        elif file_path.suffix.lower() == '.csv':
            return BatchLoader.load_csv(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {file_path.suffix}")

    @staticmethod
    def load_excel(file_path):
        """
        加载 Excel 文件

        Excel 格式要求：
        | case_id | name           | instruction                    |
        |---------|----------------|--------------------------------|
        | 1       | KOL跟评闭环    | 用一个引流号对@elonmusk...    |
        | 2       | 热点新闻蹭热度 | 用一个引流号找一个符合...     |
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        wb = load_workbook(file_path)
        ws = wb.active

        loops = []
        headers = {}

        # 列名映射（支持中文列名）
        column_aliases = {
            'case_id': ['case_id', '编号', '闭环编号', 'id'],
            'name': ['name', '闭环名称', '名称', 'loop_name'],
            'instruction': ['instruction', '指令', '闭环指令', 'command'],
        }

        # 第一行是表头
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value] = col_idx

        def get_col_idx(field):
            """根据别名查找列索引"""
            for alias in column_aliases.get(field, []):
                if alias in headers:
                    return headers[alias]
            return None

        col_case_id = get_col_idx('case_id') or 1
        col_name = get_col_idx('name') or 2
        col_instruction = get_col_idx('instruction') or 3

        # 从第二行开始读取
        for row in ws.iter_rows(min_row=2, values_only=False):
            case_id = row[col_case_id - 1].value

            if not case_id:
                continue

            name = row[col_name - 1].value
            instruction = row[col_instruction - 1].value

            if not name or not instruction:
                print(f"⚠️  跳过不完整的行: case_id={case_id}")
                continue

            loops.append({
                'case_id': case_id,
                'name': name,
                'instruction': instruction
            })

        return loops

    @staticmethod
    def load_json(file_path):
        """
        加载 JSON 文件

        JSON 格式要求：
        {
          "loops": [
            {"case_id": 1, "name": "KOL跟评闭环", "instruction": "..."},
            {"case_id": 2, "name": "热点新闻蹭热度", "instruction": "..."}
          ]
        }
        """
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        loops = data.get('loops', [])

        # 验证格式
        for loop in loops:
            if not all(k in loop for k in ['case_id', 'name', 'instruction']):
                raise ValueError(f"JSON 格式错误: 缺少必要字段 {loop}")

        return loops

    @staticmethod
    def load_csv(file_path):
        """
        加载 CSV 文件

        CSV 格式要求：
        case_id,name,instruction
        1,KOL跟评闭环,用一个引流号对@elonmusk...
        2,热点新闻蹭热度,用一个引流号找一个符合...
        """
        loops = []

        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)

            for row in reader:
                if not row.get('case_id'):
                    continue

                loops.append({
                    'case_id': row['case_id'],
                    'name': row['name'],
                    'instruction': row['instruction']
                })

        return loops
